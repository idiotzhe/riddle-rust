import random

from flask import Blueprint, request, jsonify, render_template
from web_model import db, User, Riddle, GuessRecord, Activity
from sqlalchemy import func, desc
from datetime import datetime
from openpyxl import load_workbook,Workbook

admin_bp = Blueprint('admin', __name__, url_prefix='/pro-api')

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}


def allowed_file(filename):
    """检查文件扩展名是否合法"""
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS



@admin_bp.route('/index', methods=['GET'])
def admin_index():
    return render_template('admin/index.html')


# ----------------------------
# 1. 用户管理
# ----------------------------
@admin_bp.route('/users', methods=['GET'])
def get_users():
    # pageSize: 每页大小 (默认 10)
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('pageSize', 10, type=int)
    keyword = request.args.get('keyword')
    query = User.query
    if keyword:
        query = query.filter(User.username.like(f'%{keyword}%'))
    # 2. 执行分页查询
    # SQLAlchemy 的 paginate 方法中参数依然是 per_page，我们将 page_size 传进去
    pagination = query.paginate(page=page, per_page=page_size, error_out=False)
    # 3. 组装返回格式
    return jsonify({
        "code": 200,
        "message": "success",
        "data": {
            "total": pagination.total,  # 总记录数
            "page": pagination.page,  # 当前页码
            "totalPages": pagination.pages,  # 总页数 (重命名为 totalPages)
            "list": [u.to_dict() for u in pagination.items]  # 数据列表 (重命名为 list)
        }
    })


@admin_bp.route('/user/<int:user_id>', methods=['DELETE'])
def delete_user(user_id):
    user = User.query.get(user_id)
    if user:
        db.session.delete(user)
        db.session.commit()
        return jsonify({'msg': '删除成功'})
    return jsonify({'error': '用户不存在'}), 404


# ----------------------------
# 2. 灯谜管理 (逻辑更新)
# ----------------------------
@admin_bp.route('/riddles', methods=['GET'])
def get_riddles():
    # pageSize: 每页大小 (默认 10)
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('pageSize', 10, type=int)
    keyword = request.args.get('keyword')
    query = Riddle.query
    if keyword:
        query = query.filter(Riddle.question.like(f'%{keyword}%'))

    # 2. 执行分页查询
    # SQLAlchemy 的 paginate 方法中参数依然是 per_page，我们将 page_size 传进去
    pagination = query.paginate(page=page, per_page=page_size, error_out=False)
    # 3. 组装返回格式
    return jsonify({
        "code": 200,
        "message": "success",
        "data": {
            "total": pagination.total,  # 总记录数
            "page": pagination.page,  # 当前页码
            "totalPages": pagination.pages,  # 总页数 (重命名为 totalPages)
            "list": [u.to_dict() for u in pagination.items]  # 数据列表 (重命名为 list)
        }
    })


@admin_bp.route('/riddles', methods=['POST'])
def add_riddle():
    data = request.json
    if 'id' in data and data.get("id") is not None:  # 修改
        riddle = Riddle.query.get(data['id'])
        if not riddle: return jsonify({'error': '不存在'}), 404
        riddle.question = data.get('question', riddle.question)
        riddle.answer = data.get('answer', riddle.answer)
        riddle.remark = data.get('remark', riddle.remark)

        # 处理选项修改
        if 'options' in data:
            riddle.options = data['options']

        # 允许管理员重置题目状态
        if 'reset_status' in data and data['reset_status'] is True:
            riddle.is_solved = False
            riddle.solver_id = None

    else:  # 新增
        riddle = Riddle(
            question=data['question'],
            answer=data['answer'],
            remark=data['remark']
        )
        # 设置选项
        riddle.options = data.get('options', [])
        db.session.add(riddle)

    db.session.commit()
    return jsonify(riddle.to_dict())


@admin_bp.route('/riddles/import', methods=['POST'])
def import_riddles_api():
    if 'file' not in request.files:
        return jsonify({'code': 400, 'msg': '未上传文件'})
    file = request.files['file']
    if file.filename == '' or not allowed_file(file.filename):
        return jsonify({'code': 400, 'msg': '文件格式不支持'})
    try:
        # 加载 Excel
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        # 获取表头并建立索引映射
        headers = [cell.value for cell in ws[1]]
        # 灵活匹配列名（请确保 Excel 表头与此处一致）
        try:
            q_idx = headers.index('灯谜题目')
            r_idx = headers.index('描述') if '描述' in headers else None
            ans_idx = headers.index('正确答案')
            # 动态获取所有以"选项"开头的列
            option_indices = [i for i, h in enumerate(headers) if h and '选项' in str(h)]
        except ValueError as e:
            return jsonify({'code': 400, 'msg': f'Excel缺少必要列: {str(e)}'})
        success_count = 0
        # 遍历数据行（从第二行开始）
        for row in ws.iter_rows(min_row=2, values_only=True):
            question = row[q_idx]
            answer = row[ans_idx]
            # 必填校验：题目和答案不能为空
            if not question or not answer:
                continue
            # 提取所有选项并过滤掉空值
            options_list = []
            for idx in option_indices:
                if row[idx] is not None:
                    options_list.append(str(row[idx]).strip())
            # 备注/描述
            remark = str(row[r_idx]).strip() if r_idx is not None and row[r_idx] else ""
            # 创建灯谜对象
            # 注意：这里直接给 options 赋值，会触发模型中的 @options.setter 自动转为 JSON
            random.shuffle(options_list)
            riddle = Riddle(
                question=str(question).strip(),
                remark=remark,
                answer=str(answer).strip(),
                options=options_list,  # 触发 setter
                is_solved=False
            )
            db.session.add(riddle)
            success_count += 1
        db.session.commit()
        return jsonify({
            'code': 200,
            'msg': f'灯谜导入成功: 共 {success_count} 条'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'msg': f'服务器内部错误: {str(e)}'})


@admin_bp.route('/riddles/<int:riddle_id>', methods=['DELETE'])
def delete_riddles(riddle_id):
    riddle = Riddle.query.get(riddle_id)
    if riddle:
        db.session.delete(riddle)
        db.session.commit()
        return jsonify({'msg': '删除成功'})
    return jsonify({'error': '用户不存在'}), 404


@admin_bp.route('/riddles/import', methods=['POST'])
def import_riddles():
    """批量导入"""
    data = request.json
    if not isinstance(data, list):
        return jsonify({'error': '数据格式必须为列表'}), 400

    for item in data:
        riddle = Riddle(
            question=item['question'],
            answer=item['answer']
        )
        # 导入时同步处理选项字段
        riddle.options = item.get('options', [])
        db.session.add(riddle)

    db.session.commit()
    return jsonify({'msg': '导入成功'})


@admin_bp.route('/riddle/<int:r_id>', methods=['DELETE'])
def delete_riddle(r_id):
    riddle = Riddle.query.get(r_id)
    if riddle:
        # 级联删除相关的记录，防止报错
        GuessRecord.query.filter_by(riddle_id=r_id).delete()
        db.session.delete(riddle)
        db.session.commit()
        return jsonify({'msg': '删除成功'})
    return jsonify({'error': '不存在'}), 404


# ----------------------------
# 2. 灯谜管理 (逻辑更新)
# ----------------------------
@admin_bp.route('/leaderboard', methods=['GET'])
def get_leaderboard():
    # pageSize: 每页大小 (默认 10)
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('pageSize', 10, type=int)
    keyword = request.args.get('keyword')
    query = GuessRecord.query
    if keyword:
        query = query.join(User, User.id == GuessRecord.user_id)
        query = query.filter(User.username.like(f'%{keyword}%'))
    # 2. 执行分页查询
    # SQLAlchemy 的 paginate 方法中参数依然是 per_page，我们将 page_size 传进去
    pagination = query.paginate(page=page, per_page=page_size, error_out=False)
    # 3. 组装返回格式
    return jsonify({
        "code": 200,
        "message": "success",
        "data": {
            "total": pagination.total,  # 总记录数
            "page": pagination.page,  # 当前页码
            "totalPages": pagination.pages,  # 总页数 (重命名为 totalPages)
            "list": [u.to_dict() for u in pagination.items]  # 数据列表 (重命名为 list)
        }
    })


# ----------------------------
# 4. 活动设置
# ----------------------------
@admin_bp.route('/activity', methods=['POST'])
def set_activity():
    data = request.json
    act = Activity.query.first()
    if not act or act is None:
        act = Activity()
        db.session.add(act)

    act.name = data.get('name', '元宵猜灯谜')
    act.start_time = datetime.strptime(data['start_time'], '%Y-%m-%d %H:%M:%S')
    act.end_time = datetime.strptime(data['end_time'], '%Y-%m-%d %H:%M:%S')

    db.session.commit()
    return jsonify(act.to_dict())


# ----------------------------
# 4. 活动设置
# ----------------------------
@admin_bp.route('/activity', methods=['GET'])
def get_activity():
    act = Activity.query.first()
    if not act or act is None:
        act = Activity()
        db.session.add(act)
    return jsonify(act.to_dict())
